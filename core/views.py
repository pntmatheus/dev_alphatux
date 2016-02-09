import os

import re

from django.http import HttpResponse

from .models import Pessoa

from openpyxl import load_workbook
import paramiko

"""
def index(request):
    return HttpResponse("Olá Mundo!!!")
"""


def executa_comando(comando, ssh_session, v5=False):

    if v5 is False:
        mikrotik = ":do { :put [%s];} on-error={ :put \"###ERRO###\"; }" % comando
        stdin, stdout, stderr = ssh_session.exec_command(mikrotik)
        teste2 = stdout.read()
        teste2 = teste2.decode("utf-8")

        if "###ERRO###" in teste2:
            stdin, stdout, stderr = ssh_session.exec_command(comando)
            teste = stdout.read()
            teste = teste.decode("utf-8")
            retorno = "FALHOU!!!!     ----------->     " + mikrotik + " ----->> " + teste
        else:
            retorno = "OK! ---------->" + mikrotik + "----->" + teste2
    else:
        stdin, stdout, stderr = ssh_session.exec_command(comando)
        teste2 = stdout.read()
        #teste2 = teste2.decode("utf-8")
        retorno = "%s ---> %s" % (comando, teste2)

    return retorno

#Achei interessante separar em duas funções para melhor trabalhar no "find" do Mikrotik
def gera_codigo_comentario(id):
    return "##%s##" % id


def faz_cometario(nome, id):
    return "\"%s --- %s\"" % (gera_codigo_comentario(id), nome)

def get_ultimo_ip():
    wb = load_workbook(filename='clientesTerezinha.xlsx', read_only=True)
    ws = wb['Plan1']
    ip = ws["D1"].value
    return ip

def set_ultimo_ip(ip):
    wb = load_workbook(filename='clientesTerezinha.xlsx', read_only=False)
    ws = wb['Plan1']
    ip = ws["D1"].value
    ws["D1"].value = ip
    wb.save(filename='clientesTerezinha.xlsx')

def gera_proximos_ips(ip_cliente):
    ips = []

    part1, part2, part3, part4 = ip_cliente.split('.')
    sufixo = "%s.%s" % (part1, part2)
    contador = int(part3)
    cliente = int(part4)
    rede = cliente - 2
    gateway = cliente - 1
    if contador == 200:
        contador = 20
        gateway += 4
        rede += 4
        cliente += 4
    else:
        contador += 1

    mascara = "/30"

    ip_teste = "%s.%s.%s" % (sufixo, contador, cliente)
    return ip_teste

'''
it accepts 12 hex digits with either : or - as separators between pairs (but the separator must be uniform... either all separators are : or are all -).

This is the explanation:

    ## [0-9a-f] means an hexadecimal digit
    ## {2} means that we want two of them
    ## [-:] means either a dash or a colon. Note that the dash as first char doesn't mean a range but only means itself. This atom is enclosed in parenthesis so it can be reused later as a back reference.
    ## [0-9a-f]{2} is another pair of hexadecimal digits
    ## \\1 this means that we want to match the same expression that we matched \"before as separator. This is what guarantees uniformity. Note that the regexp syntax is \1 but I'm using a regular string so backslash must be escaped by doubling it.
    ## [0-9a-f]{2} another pair of hex digits
    ## {4} the previous parenthesized block must be repeated exactly 4 times, giving a total of 6 pairs of digits: <pair> <sep> <pair> ( <same-sep> <pair> ) * 4
    ## $ The string must end right after them

'''


def valida_mac(mac):
    if re.match("[0-9a-f]{2}([-:])[0-9a-f]{2}(\\1[0-9a-f]{2}){4}$", mac.lower()):
        mac = mac.upper()
        return mac
    else:
        return "FALHOU PQ O %s NAO EH UM MAC VALIDO!!!!" % mac



def valida_comando(ssh_session, comandos):
    retorno = []
    for comando in comandos:
        mikrotik = ":do { :put [%s];} on-error={ :put \"###ERRO###\"; }" % comando
        stdin, stdout, stderr = ssh_session.exec_command(mikrotik)
        teste2 = stdout.read()
        teste2 = teste2.decode("utf-8")

        if "###ERRO###" in teste2:
            stdin, stdout, stderr = ssh_session.exec_command(comando)
            teste = stdout.read()
            teste = teste.decode("utf-8")
            retorno.append("FALHOU!!!!     ----------->     " + mikrotik + " ----->> " + teste + "<br>")
        else:
            retorno.append("OK! ---------->" + mikrotik + "----->" + teste2 + " <br>")

    return retorno



'''
def valida_comando(ssh_session, comandos):
    retorno = []

    for comando in comandos:
        stdin, stdout, stderr = ssh_session.exec_command(comando)
        retorno.append("%s" % (comando))

    return retorno '''

def add_rota(id, nome, dst_address, gateway):
    comment = faz_cometario(nome, id)
    comando = "/ip route add dst-address=%s gateway=%s comment=%s" % (dst_address, gateway, comment)
    return comando

def add_plano(nome, download, upload, comandos):
    plano = "/ip hotspot user profile add idle-timeout=none name=%s rate-limit=%s/%s shared-users=1 status-autorefresh=1m transparent-proxy=no add-mac-cookie=no" % (
        nome, upload, download)
    plano_bloqueado = "/ip hotspot user profile add advertise=yes advertise-interval=0s advertise-timeout=immediately advertise-url=bloqueado.html idle-timeout=none name=bloqueado-%s open-status-page=always shared-users=1 status-autorefresh=1m transparent-proxy=yes add-mac-cookie=no" % (
        nome)
    comandos.append(plano)
    comandos.append(plano_bloqueado)
    return comandos


def add_ip_gateway(nome, ip, id):
    comment = faz_cometario(nome, id)
    comando = "/ip address add address=%s interface=bridge-clientes comment=%s" % (ip, comment)
    return comando


def add_dhcp_network(nome, ip_rede, ip_gateway, ip_dns, id):
    comment = faz_cometario(nome, id)
    comando = "/ip dhcp-server network add address=%s comment=%s dns-server=%s gateway=%s" % (
        ip_rede, comment, ip_dns, ip_gateway)
    return comando


def add_dhcp_lease(nome, ip_cliente, mac, id):
    comment = faz_cometario(nome, id)
    mac = valida_mac(mac)
    comando = "/ip dhcp-server lease add address=%s comment=%s mac-address=%s" % (ip_cliente, comment, mac)
    return comando


def add_hotspot_user(nome, mac, plano, id):
    comment = faz_cometario(nome, id)
    mac = valida_mac(mac)
    comando = "/ip hotspot user add comment=%s mac-address=%s name=%s profile=%s" % (comment, mac, mac, plano)
    return comando


def add_hotspot_ip_binding(nome, ip_cliente, mac, id):
    comandos = []
    comment = faz_cometario(nome, id)
    mac = valida_mac(mac)
    comandos.append(
        "/ip hotspot ip-binding add address=%s comment=%s mac-address=%s type=regular" % (ip_cliente, comment, mac))
    comandos.append(
        "/ip hotspot ip-binding move destination=[find comment=\"BLOQUEAR RESTANTE\"] numbers=[find comment=%s];" % comment)
    return comandos

def add_access_list(id, nome, mac):
    comment = faz_cometario(nome, id)
    mac = valida_mac(mac)
    comando = "/interface wireless access-list add mac-address=%s comment=%s forwarding=no" % (mac, comment)

    return comando

def add_cliente(id, nome, sufixo, mac, plano, v5=False, frequencia24ghz=False):

    versao = v5
    frequencia = frequencia24ghz
    retorno = []

    path = os.path.join(os.environ['HOME'], '.ssh', 'id_dsa')
    pkey = paramiko.DSSKey.from_private_key_file(path, password=None)

    # Hotspot Z3
    ssh_hotz3 = paramiko.SSHClient()
    ssh_hotz3.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_hotz3.connect('192.168.26.1', username='adminDjango', pkey=pkey, port=222)

    # Roteador Z3
    ssh_rotz3 = paramiko.SSHClient()
    ssh_rotz3.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_rotz3.connect('192.168.2.1', username='adminDjango', pkey=pkey, port=222)

    # Roteador Terezinha
    ssh_rot_terezinha = paramiko.SSHClient()
    ssh_rot_terezinha.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_rot_terezinha.connect('10.2.2.1', username='adminDjango', pkey=pkey, port=222)


    if (frequencia24ghz):
        # Alphatux Turbo 2
        ssh_alphatuxturbo2 = paramiko.SSHClient()
        ssh_alphatuxturbo2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatuxturbo2.connect('192.168.15.7', username='adminDjango', pkey=pkey, port=222)

        # Alphatux Cedro --- Comentado ate atualizar a RB para versao 5 ou 6
        # ssh_alphatuxcedro = paramiko.SSHClient()
        # ssh_alphatuxcedro.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # ssh_alphatuxcedro.connect('192.168.15.8', username='adminDjango', pkey=pkey, port=222)

        # Alphatux
        ssh_alphatux = paramiko.SSHClient()
        ssh_alphatux.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatux.connect('192.168.15.10', username='adminDjango', pkey=pkey, port=222)

        # Alphatuxcd2
        ssh_alphatuxcd2 = paramiko.SSHClient()
        ssh_alphatuxcd2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatuxcd2.connect('192.168.15.12', username='adminDjango', pkey=pkey, port=222)

        # Alphatuxcs1
        ssh_alphatuxcs1 = paramiko.SSHClient()
        ssh_alphatuxcs1.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatuxcs1.connect('192.168.15.14', username='adminDjango', pkey=pkey, port=222)

        retorno.append("-----------------------AlphatuxTurbo2-------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxturbo2, versao) + "<br>")

        # retorno.append("------------------------AlphatuxCedro-------------------------------<br>")
        # retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxcedro, versao) + "<br>")

        retorno.append("------------------------Alphatux-------------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatux, versao) + "<br>")

        retorno.append("------------------------AlphatuxCedro2-------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxcd2, versao) + "<br>")

        retorno.append("------------------------AlphatuxCASA1-------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxcs1, versao) + "<br>")

    else:
        # Alphatuxz31
        ssh_alphatuxz31 = paramiko.SSHClient()
        ssh_alphatuxz31.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatuxz31.connect('192.168.15.20', username='adminDjango', pkey=pkey, port=222)

        # Alphatuxz32
        ssh_alphatuxz32 = paramiko.SSHClient()
        ssh_alphatuxz32.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_alphatuxz32.connect('192.168.15.21', username='adminDjango', pkey=pkey, port=222)

        retorno.append("-----------------------AlphatuxZ31-------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxz31, versao) + "<br>")

        retorno.append("------------------------AlphatuxZ32-------------------------------<br>")
        retorno.append(executa_comando(add_access_list(id, nome, mac), ssh_alphatuxz32, versao) + "<br>")



    rede = 4
    gateway = 5
    cliente = 6
    mascara = "/30"

    ip_gateway = "%s.%s" % (sufixo, str(gateway))
    ip_rede = "%s.%s%s" % (sufixo, str(rede), mascara)
    ip_dns = "192.168.2.1"
    ip_cliente = "%s.%s" % (sufixo, str(cliente))

    # Comando avulso para driblar o erro inicial do mikrotik
    retorno.append(executa_comando("/interface get number=0", ssh_hotz3, versao) + "<br>")
    retorno.append(executa_comando(add_ip_gateway(nome, ip_gateway + mascara, id), ssh_hotz3, versao) + "<br>")
    retorno.append(executa_comando(add_dhcp_network(nome, ip_rede, ip_gateway, ip_dns, id), ssh_hotz3, versao) + "<br>")
    retorno.append(executa_comando(add_dhcp_lease(nome, ip_cliente, mac, id), ssh_hotz3, versao) + "<br>")
    retorno.append(executa_comando(add_hotspot_user(nome, mac, plano, id), ssh_hotz3, versao) + "<br>")

    for comando in add_hotspot_ip_binding(nome, ip_cliente, mac, id):
        retorno.append(executa_comando(comando, ssh_hotz3, versao) + "<br>")

    retorno.append("---------------------------------------------------------------<br>")
    retorno.append(executa_comando(add_rota(id, nome, ip_rede, "192.168.2.2"), ssh_rotz3, versao) + "<br>")

    retorno.append("---------------------------------------------------------------<br>")
    retorno.append(executa_comando(add_rota(id, nome, ip_rede, "10.2.2.2"), ssh_rot_terezinha, versao) + "<br>")

    return retorno


## ALTERA O IP NO /IP ADDRESS - INFORMAR ***SEMPRE*** IP+MASCARA (/30, /24, ...)
def altera_ip_ip_address(id, ip):
    comando = "/ip address set [/ip address find comment~\"%s\"] address=%s" % (gera_codigo_comentario(id), ip)
    return comando


## ALTERA O NOME DO CLIENTE NO COMENTARIO DO IP ADDRESS
def altera_nome_ip_address(id, nome):
    comando = "/ip address set [/ip address find comment~\"%s\"] comment=%s" % (
        gera_codigo_comentario(id), faz_cometario(nome, id))
    return comando


## ALTERA OS IPS DE REDE E GATEWAY DO DHCP NETWORKS, PRECISA SER MELHORA A QUESTAO DE VALIDAR OS IPS
###IP DA REDE DEVE SER ACOMPANHADO PELO BIT DE MASCARA
def altera_ip_dhcp_networks(id, ip_rede, ip_gateway):
    comando = "/ip dhcp-server network set [find comment~\"%s\"] address=%s gateway=%s" % (
        gera_codigo_comentario(id), ip_rede, ip_gateway)
    return comando


def altera_nome_dhcp_networks(id, nome):
    comando = "/ip dhcp-server network set [find comment~\"%s\"] comment=%s;" % (
        gera_codigo_comentario(id), faz_cometario(nome, id))
    return comando


def altera_ip_dhcp_lease(id, ip):
    comando = "/ip dhcp-server lease set [find comment~\"%s\"] address=%s" % (gera_codigo_comentario(id), ip)
    return comando


def altera_mac_dhcp_lease(id, mac):
    comando = "/ip dhcp-server lease set [find comment~\"%s\"] mac-address=\"%s\"" % (gera_codigo_comentario(id), mac)
    return comando


def altera_nome_dhcp_lease(id, nome):
    comando = "/ip dhcp-server lease set [find comment~\"%s\"] comment=%s;" % (
        gera_codigo_comentario(id), faz_cometario(nome, id))
    return comando


def altera_mac_hotspot_user(id, mac):
    comando = "/ip hotspot user set [find comment~\"%s\"] name=\"%s\" mac-address=\"%s\"" % (
        gera_codigo_comentario(id), mac, mac)
    return comando


def altera_profile_hotspot_user(id, profile):
    comando = "/ip hotspot user set [find comment~\"%s\"] profile=\"%s\"" % (gera_codigo_comentario(id), profile)
    return comando


def altera_nome_hotspot_user(id, nome):
    comando = "/ip hotspot user set [find comment~\"%s\"] comment=%s;" % (
        gera_codigo_comentario(id), faz_cometario(nome, id))
    return comando


def altera_ip_hotspot_bindings(id, ip):
    comando = "/ip hotspot ip-binding set [find comment~\"%s\"] address=%s" % (gera_codigo_comentario(id), ip)
    return comando


def busca_existencia(comando):
    verifica = ":if ([:len [%s]] > 0) do={:put \"Found\";} else={:put \"Not Found\";};" % comando
    return verifica


def index(request):



    backup = False
    comandos = []

    if (backup):
        # CONFIGURACOES INICIAIS
        wb = load_workbook(filename='clientesTerezinha.xlsx', read_only=True)
        ws = wb['Plan1']

        clientes = []
        macs = []
        ids = []

        for p in ws["A1":"C76"]:
            ids.append(p[0].value)
            clientes.append(p[1].value)
            macs.append(p[2].value)


        ##Criar BRIDGE CLIENTES
        comandos.append(
            "/interface bridge add comment=\"Bridge dos Clientes\" name=bridge-clientes admin-mac=[/interface get ether1 mac-address] auto-mac=no")
        ##SETAR IP NA BRIDGE
        comandos.append(
            "/ip address add address=192.168.10.1/24 comment=\"IP Inicial da BRIDGE CLIENTES\" interface=bridge-clientes")
        ##Adicionar INTERFACE 1 NA BRIDGE
        comandos.append("/interface bridge port add bridge=bridge-clientes interface=ether1")
        ##Criar DHCP SERVER
        comandos.append(
            "/ip dhcp-server add add-arp=yes lease-time=3d address-pool=static-only interface=bridge-clientes name=\"AlphatuxZ3\" disabled=no")
        ##Adicionar PROFILE HOTSPOT
        comandos.append(
            "/ip hotspot profile add dns-name=\"\" hotspot-address=0.0.0.0 html-directory=hotspot http-proxy=0.0.0.0:0 login-by=mac mac-auth-password=\"\" name=alphatux-z3 rate-limit=\"\" smtp-server=0.0.0.0 use-radius=no")
        ##Adicionar Planos no HOTSPOT USER PROFILE
        ###Basico
        comandos = add_plano("basico", "256k", "156k", comandos)
        ###Intermediario
        comandos = add_plano("intermediario", "500k", "300k", comandos)
        ###1M
        comandos = add_plano("1M", "1000k", "400k", comandos)
        ###2M
        comandos = add_plano("2M", "2000k", "400k", comandos)
        ###Matheus
        comandos = add_plano("matheus", "14500k", "2500k", comandos)
        ###Ultra
        comandos = add_plano("ultra", "3000k", "100k", comandos)
        ##Adicionar REGRA IP BINDING
        comandos.append("/ip hotspot ip-binding add address=0.0.0.0/0 comment=\"BLOQUEAR RESTANTE\" type=blocked")


        # ADD CLIENTES
        sufixo = "172.21"
        contador = 20
        rede = 0
        gateway = 1
        cliente = 2
        mascara = "/30"
        for (nome_cliente, mac, id_cliente) in zip(clientes, macs, ids):
            if contador > 200:
                contador = 20
                gateway += 4
                rede += 4
                cliente += 4

            ip_gateway = "%s.%s.%s" % (sufixo, str(contador), str(gateway))
            ip_rede = "%s.%s.%s%s" % (sufixo, str(contador), str(rede), mascara)
            ip_dns = "192.168.2.1"
            ip_cliente = "%s.%s.%s" % (sufixo, str(contador), str(cliente))

            ###Adicionar IPS GATEWAY-CLIENTE
            comandos.append(add_ip_gateway(nome_cliente, ip_gateway + mascara, id_cliente))
            ###Adicionar DHCP Networks
            comandos.append(add_dhcp_network(nome_cliente, ip_rede, ip_gateway, ip_dns, id_cliente))
            ###Adicionar DHCP Lease
            comandos.append(add_dhcp_lease(nome_cliente, ip_cliente, mac, id_cliente))
            ###Adicionar HOTSPOT USER
            comandos.append(add_hotspot_user(nome_cliente, mac, "1M", id_cliente))
            ###Adicionar HOTSPOT IP BINDING
            comandos.extend(add_hotspot_ip_binding(nome_cliente, ip_cliente, mac, id_cliente))
            contador += 1

        ##ADD HOTSPOT --- COLOQUEI AQUI PQ LOGO QUE SE CRIA O HOTSPOT O MIKROTIK BLOQUEIA AS ENTRADAS PELO FIREWALL
        comandos.append("/ip hotspot ip-binding add address=192.168.56.1 type=bypassed comment=\"ServidorDjango\"")
        comandos.append("/ip hotspot ip-binding move destination=0 numbers=[find comment~\"ServidorDjango\"];")
        ##Adicionar HOTSPOT
        comandos.append(
            "/ip hotspot add disabled=no idle-timeout=none interface=bridge-clientes keepalive-timeout=none name=Alphatux profile=alphatux-z3")

        path = os.path.join(os.environ['HOME'], '.ssh', 'id_dsa')
        pkey = paramiko.DSSKey.from_private_key_file(path, password=None)
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect('192.168.56.10', username='admin', pkey=pkey)

        retorno = []

        for comando in comandos:
            retorno.append(executa_comando(comando, ssh) + "<br>")

    teste = False

    if (teste):
        mamae = 0
        # tt#ALTERAR CLIENTES
        # comandos.append(":if ([:len [:find \"abcd\" \"x\"]] > 0) do={:put \"Found\";} else={:put \"Not Found\";};")
        ##teste altera_ip_ip_address
        # comandos.append(altera_ip_ip_address(4, "192.168.23.5/24"))
        ##teste altera_nome_ip_address
        # comandos.append(altera_nome_ip_address(4, "Testando..."))
        ##teste altera_ip_dhcp_networks
        # comandos.append(altera_ip_dhcp_networks(5, "192.168.25.8/30", "192.168.25.9"))
        # teste altera_nome_dhcp_networks
        # comandos.append(altera_nome_dhcp_networks(75, "Matheus2"))
        # teste altera_ip_dhcp_lease
        # comandos.append(altera_ip_dhcp_lease(5, "192.168.24.6"))
        # teste altera_ip_dhcp_lease
        # comandos.append(altera_mac_dhcp_lease(6, valida_mac("00:0c:29:de:a3:7a")))
        # teste altera_nome_dhcp_lease
        # comandos.append(altera_nome_dhcp_lease(7, "Matheus Testado..."))
        # teste altera_mac_hotspot_user
        # comandos.append(altera_mac_hotspot_user(4, valida_mac("00:0c:29:de:a3:7a")))
        # teste altera_profile_hotspot_user
        # comandos.append(altera_profile_hotspot_user(9, "2M"))
        # teste altera_nome_hotspot_user
        # comandos.append(altera_nome_hotspot_user(5, "Amanda Teste"))
        # teste altera_ip_hotspot_binding
        # comandos.append(altera_ip_hotspot_bindings(7, "192.168.26.6"))
        # comandos.append("/ip address print")

    #resultado = add_cliente(4, "Analu", "192.168.137", "58:10:8C:04:E3:38")



    comando = "/ip address print"
    return HttpResponse(add_cliente(18,"Bruno Fagundes","192.168.152","E4:8D:8C:C2:D9:97","1M",True, frequencia24ghz=False))
### 1 4C:5E:0C:FF:7B:39 Miguel Irigon
### 2 4C:5E:0C:EA:7B:53 Lucas Pereira
### 3 4C:5E:0C:A8:B2:BC Sabrina Ramires
### 4 4C:5E:0C:0D:C8:1F Diene Batista Rodrigues
def teste (request):
    return HttpResponse("teste")

